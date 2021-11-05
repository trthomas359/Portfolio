from openpyxl import Workbook, load_workbook #biblioteca que trabalha com excel
import os #biblioteca de manipulacao de arquivos


#procura condicionais de erro para a pasta

def compile_workbooks(workbooks_path, final_filename):
    if not isinstance(workbooks_path, str):
        raise TypeError("caminho da pasta tem que ser string")
    
    if not isinstance(final_filename, str):
        raise TypeError("nome do arquivo tem que ser string")

    if not os.path.exists(workbooks_path):
        raise NotADirectoryError("workbook_path indicado não existe")

    if not final_filename.endswith(".xlsx"):
        raise ValueError('nome final do arquivo deve ter extensao ".xlsx"')
        
    if final_filename in os.listdir(workbooks_path):
        raise ValueError(f'ja tem arquivo {final_filename} no {workbooks_path}. '
                         f'mova o arquivo ou mude o nome do final_filename')
    

#procura os arquivos com final xlsx

    wbs = []
    for file in os.listdir(workbooks_path): 
        if not file.startswith("~$") and file.endswith(".xlsx"):
            wb = load_workbook(os.path.join(workbooks_path, file))
            wbs.append(wb)
    
    final_wb = Workbook()
    final_ws = final_wb.worksheets[0]

    wb1 = wbs[0]
    ws1 = wb1.worksheets[0] 
 
#adiciona os valores em loop na nova plan


    for j in range(1, ws1.max_column+1):
        final_ws.cell(row=1, column=j).value = ws1.cell(row=1, column=j).value

    current_row = 2

    for wb in wbs:
        for ws in wb.worksheets:
            mr = ws.max_row 
            mc = ws.max_column 

            for i in range (2, mr + 1): 
                for j in range (1, mc + 1): 
                    current_cell = ws.cell(row = i, column = j) 
                    final_ws.cell(row = current_row, column = j).value = current_cell.value

                current_row += 1

    final_wb.save(os.path.join(workbooks_path, final_filename)) #salvar

#indica o caminho, compila e salva com o nome final

if __name__ == '__main__':
    compile_workbooks(os.path.join(os.getcwd(), "temp fundos"), "final.xlsx") 
    #mudar o caminho para a pasta do dia
